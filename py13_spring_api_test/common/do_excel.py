import openpyxl
from openpyxl import load_workbook
class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.method = None
        self.url = None
        self.data = None
        self.expected = None

class DoExcel: #定义一个从excel读取数据的类
    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook =load_workbook(file_name)

    def read_data(self,sheet_name): #从excel读取数据
        self.sheet_name = sheet_name
        sheet = self.workbook[sheet_name]
        cases =[]
        for item in range(2,sheet.max_row+1):
            row_case = Case()
            row_case.case_id = sheet.cell(row =item, column = 1).value
            row_case.title = sheet.cell(row =item, column = 2).value
            row_case.method = sheet.cell(row =item, column = 3).value
            row_case.url = sheet.cell(row =item, column = 4).value
            row_case.data = sheet.cell(row =item, column = 5).value
            row_case.expected = sheet.cell(row =item, column = 6).value
            cases.append(row_case)
        return cases
    def write_back(self, row, actual, result): #写回数据到excel
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row=row, column =7,value=actual)
        sheet.cell(row=row, column=8, value=result)
        self.workbook.save(self.file_name)

if __name__ == '__main__':
    from common import constant
    do_execl =DoExcel(constant.case_dir)
    do_execl.read_data('register')

